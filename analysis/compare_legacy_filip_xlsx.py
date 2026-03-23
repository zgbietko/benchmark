#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
from datetime import datetime
import json
import math
import os
from pathlib import Path
import re
import sys
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
PLOTS_ROOT = ROOT / "analysis" / "plots"
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

_mpl_cfg = ROOT / ".cache" / "matplotlib"
_mpl_cfg.mkdir(parents=True, exist_ok=True)
os.environ.setdefault("MPLCONFIGDIR", str(_mpl_cfg))

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # type: ignore
import numpy as np  # type: ignore

try:
    from openpyxl import load_workbook  # type: ignore
except Exception as exc:  # pragma: no cover - dependency guard
    raise SystemExit(
        "openpyxl is required for XLSX import. Install it with: python3 -m pip install openpyxl"
    ) from exc


VARIANT_ORDER = ["qss", "sqs", "ssq"]
OPTION_LABEL = "Options (9 bits, Filip order)"
LEGACY_OPTION_COLUMNS = [
    "-D COAL_READ",
    "-D COAL_WRITE",
    "-D COMPUTE_ALL_SHAPE_FUN_DER",
    "-D USE_WORKSPACE_FOR_PDE_COEFF",
    "-D USE_WORKSPACE_FOR_GEO_DATA",
    "-D USE_WORKSPACE_FOR_SHAPE_FUN",
    "-D USE_WORKSPACE_FOR_STIFF_MAT",
    "-D WORKSPACE_PADDING=0",
    "-D WORKSPACE_PADDING=1",
]


def _safe_float(value: Any) -> float:
    try:
        if value is None:
            return float("nan")
        return float(value)
    except Exception:
        return float("nan")


def _safe_int(value: Any, default: int = -1) -> int:
    try:
        return int(value)
    except Exception:
        return int(default)


def _variant_title(name: str) -> str:
    return str(name).upper()


def _stacked_combo_label(bits: str) -> str:
    return "\n".join(list(str(bits)))


def _combo_fontsize(combos: list[str]) -> float:
    n_points = len(combos)
    if n_points >= 72:
        return 3.8
    if n_points >= 48:
        return 4.2
    if n_points >= 24:
        return 4.8
    return 5.4


def _setup_plot_style() -> None:
    plt.style.use("default")
    plt.rcParams.update(
        {
            "figure.facecolor": "white",
            "axes.facecolor": "white",
            "axes.grid": True,
            "grid.alpha": 0.25,
            "font.family": "serif",
            "axes.titlesize": 11,
            "axes.labelsize": 10,
            "legend.fontsize": 8,
        }
    )


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(text).lower()).strip("_") or "legacy"


def _infer_variant(path: Path) -> str:
    stem = path.stem.lower()
    for variant in VARIANT_ORDER:
        if f"_{variant}_" in stem or stem.endswith(f"_{variant}") or f"_{variant}" in stem:
            return variant
    raise ValueError(f"Cannot infer variant from filename: {path.name}")


def _infer_case_label(paths: list[Path]) -> str:
    names = [p.stem.lower() for p in paths]
    if names and all("laplace" in n and "prism" in n for n in names):
        return "Legacy Intel Xe SVM | Laplace prism"
    if names and all(re.search(r"(^|_)test(_|$)", n) and "prism" in n for n in names):
        return "Legacy Intel Xe SVM | TEST prism"
    return "Legacy Filip XLSX reference"


def _guess_n_qp(paths: list[Path], explicit: int) -> tuple[int, str]:
    if explicit > 0:
        return explicit, "explicit"
    if paths and all("prism" in p.stem.lower() for p in paths):
        return 6, "inferred_from_prism_filename"
    return 1, "default_1"


def _infer_legacy_case(path: Path) -> tuple[str, str, str]:
    stem = path.stem.lower()
    if "laplace" in stem and "prism" in stem:
        return "laplace_prism", "laplace", "prism6"
    if re.search(r"(^|_)test(_|$)", stem) and "prism" in stem:
        return "test_prism", "test", "prism6"
    return "legacy_reference", "unknown", "unknown"


def _read_legacy_xlsx(path: Path, *, n_qp: int, label: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    try:
        headers = [str(v).strip() if v is not None else "" for v in next(it)]
    except StopIteration:
        return []
    header_map = {name: idx for idx, name in enumerate(headers)}
    for name in LEGACY_OPTION_COLUMNS + ["nr_elems", "executing kernel", "internal"]:
        if name not in header_map:
            raise ValueError(f"Missing expected column '{name}' in {path.name}")

    rows: list[dict[str, Any]] = []
    variant = _infer_variant(path)
    case_name, operator_name, element_type = _infer_legacy_case(path)
    for option_index, raw in enumerate(it):
        if not raw or raw[0] is None:
            continue
        option_row = [_safe_int(raw[header_map[col]], 0) for col in LEGACY_OPTION_COLUMNS]
        combo_bits = "".join("1" if v != 0 else "0" for v in option_row)
        n_elements = max(1.0, _safe_float(raw[header_map["nr_elems"]]))
        kernel_s = _safe_float(raw[header_map["executing kernel"]])
        internal_s = _safe_float(raw[header_map["internal"]])
        input_s = _safe_float(raw[header_map.get("sending el_data_in to GPU memory", -1)]) if "sending el_data_in to GPU memory" in header_map else float("nan")
        output_s = _safe_float(raw[header_map.get("copying output buffer", -1)]) if "copying output buffer" in header_map else float("nan")
        rows.append(
            {
                "source": "legacy_xlsx",
                "source_file": path.name,
                "sheet": ws.title,
                "label": label,
                "case": case_name,
                "operator": operator_name,
                "element_type": element_type,
                "variant": variant,
                "option_index": option_index,
                "option_row": option_row,
                "combo_bits": combo_bits,
                "n_elements": n_elements,
                "n_qp": int(n_qp),
                "kernel_time_s": kernel_s,
                "internal_time_s": internal_s,
                "input_time_s": input_s,
                "output_time_s": output_s,
                "kernel_ms": kernel_s * 1e3 if math.isfinite(kernel_s) else float("nan"),
                "internal_ms": internal_s * 1e3 if math.isfinite(internal_s) else float("nan"),
                "kernel_ns_per_elem": kernel_s * 1e9 / n_elements if math.isfinite(kernel_s) else float("nan"),
                "kernel_ns_per_unit": kernel_s * 1e9 / max(1.0, n_elements * max(n_qp, 1)) if math.isfinite(kernel_s) else float("nan"),
            }
        )
    return rows


def _load_optimization_run(path: Path) -> dict[str, Any]:
    summary = json.loads((path / "summary.json").read_text(encoding="utf-8"))
    rows: list[dict[str, Any]] = []
    with (path / "evaluations.jsonl").open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                raw = json.loads(line)
            except Exception:
                continue
            cfg = raw.get("config")
            if not isinstance(cfg, dict):
                cfg = {k[4:]: v for k, v in raw.items() if k.startswith("cfg_")}
            metrics = raw.get("metrics")
            if not isinstance(metrics, dict):
                metrics = {k[7:]: v for k, v in raw.items() if k.startswith("metric_")}
            if not cfg or not metrics:
                continue
            option_row = raw.get("option_row", [])
            if not isinstance(option_row, list):
                option_row = []
            combo_bits = "".join("1" if _safe_int(v, 0) != 0 else "0" for v in option_row) if option_row else ""
            elapsed = _safe_float(metrics.get("elapsed_s_mean"))
            n_elements = max(1.0, _safe_float(metrics.get("n_elements")))
            n_qp = max(1, _safe_int(metrics.get("n_qp_effective"), _safe_int(metrics.get("n_qp_requested"), 1)))
            rows.append(
                {
                    "source": "optimization_run",
                    "label": f"{summary.get('resolved_backend', summary.get('backend', 'run'))} | {summary.get('device', 'unknown')}",
                    "run_dir": str(path),
                    "operator": str(cfg.get("operator", "")).strip().lower(),
                    "element_type": str(cfg.get("element_type", "")).strip().lower(),
                    "variant": str(cfg.get("algorithm_variant", "")).strip().lower(),
                    "option_index": _safe_int(raw.get("option_index"), -1),
                    "option_row": option_row,
                    "combo_bits": combo_bits,
                    "n_elements": n_elements,
                    "n_qp": n_qp,
                    "kernel_time_s": elapsed,
                    "kernel_ms": elapsed * 1e3 if math.isfinite(elapsed) else float("nan"),
                    "kernel_ns_per_elem": elapsed * 1e9 / n_elements if math.isfinite(elapsed) else float("nan"),
                    "kernel_ns_per_unit": elapsed * 1e9 / max(1.0, n_elements * n_qp) if math.isfinite(elapsed) else float("nan"),
                    "status": str(raw.get("status", "")),
                    "constraints_ok": _safe_int(raw.get("constraints_ok"), 0) == 1,
                }
            )
    return {"summary": summary, "rows": rows}


def _preferred_current_rows(rows: list[dict[str, Any]], operator: str) -> list[dict[str, Any]]:
    filtered = [r for r in rows if r.get("operator") == operator]
    ok = [r for r in filtered if str(r.get("status", "")) == "ok" and bool(r.get("constraints_ok"))]
    return ok if ok else filtered


def _best_by_combo(rows: list[dict[str, Any]], variant: str, metric_key: str) -> tuple[list[str], list[float]]:
    best: dict[str, float] = {}
    order_map: dict[str, int] = {}
    for row in rows:
        if row.get("variant") != variant:
            continue
        combo = str(row.get("combo_bits", "")).strip()
        if not combo:
            continue
        val = _safe_float(row.get(metric_key))
        if not math.isfinite(val):
            continue
        idx = _safe_int(row.get("option_index"), -1)
        if idx >= 0:
            prev_idx = order_map.get(combo)
            if prev_idx is None or idx < prev_idx:
                order_map[combo] = idx
        prev = best.get(combo)
        if prev is None or val < prev:
            best[combo] = val
    combos = sorted(
        best.keys(),
        key=lambda combo: (0 if combo in order_map else 1, order_map.get(combo, 10**9), combo),
    )
    return combos, [best[c] for c in combos]


def _union_combos(series_rows: list[list[dict[str, Any]]], variant: str) -> list[str]:
    combos: set[str] = set()
    order_map: dict[str, int] = {}
    for rows in series_rows:
        for row in rows:
            if row.get("variant") != variant:
                continue
            combo = str(row.get("combo_bits", "")).strip()
            if not combo:
                continue
            combos.add(combo)
            idx = _safe_int(row.get("option_index"), -1)
            if idx >= 0:
                prev_idx = order_map.get(combo)
                if prev_idx is None or idx < prev_idx:
                    order_map[combo] = idx
    return sorted(combos, key=lambda combo: (0 if combo in order_map else 1, order_map.get(combo, 10**9), combo))


def _plot_series(ax: Any, *, combos: list[str], series: list[tuple[str, dict[str, float]]], title: str, ylabel: str) -> None:
    xs = list(range(len(combos)))
    colors = ["#111111", "#2563eb", "#b45309", "#0f766e", "#7c3aed"]
    all_vals: list[float] = []
    for idx, (label, mapping) in enumerate(series):
        ys = [mapping.get(combo, float("nan")) for combo in combos]
        cleaned = [float(v) if math.isfinite(float(v)) else np.nan for v in ys]
        ax.plot(xs, cleaned, linewidth=1.5, color=colors[idx % len(colors)], label=label)
        all_vals.extend(float(v) for v in ys if math.isfinite(float(v)))
    ax.set_title(title)
    ax.set_ylabel(ylabel)
    ax.set_xticks(xs, [_stacked_combo_label(combo) for combo in combos], fontsize=_combo_fontsize(combos))
    ax.tick_params(axis="x", length=0, pad=3)
    ax.set_xlim(-0.5, len(combos) - 0.5)
    ax.grid(True, axis="y", alpha=0.28)
    if all_vals:
        robust_hi = float(np.percentile(np.array(all_vals, dtype=float), 97.0))
        max_hi = max(all_vals)
        if robust_hi > 0.0 and max_hi > robust_hi * 1.35:
            ax.set_ylim(0.0, robust_hi * 1.10)
            ax.text(0.99, 0.96, "high outliers clipped", transform=ax.transAxes, ha="right", va="top", fontsize=8, color="#475569")
    ax.legend(loc="upper center", ncol=max(1, min(3, len(series))), frameon=False)


def _write_rows_jsonl(rows: list[dict[str, Any]], out_path: Path) -> None:
    with out_path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=True) + "\n")


def _write_rows_csv(rows: list[dict[str, Any]], out_path: Path) -> None:
    fieldnames = [
        "source",
        "source_file",
        "sheet",
        "label",
        "case",
        "operator",
        "element_type",
        "variant",
        "option_index",
        "combo_bits",
        "n_elements",
        "n_qp",
        "kernel_time_s",
        "internal_time_s",
        "input_time_s",
        "output_time_s",
        "kernel_ms",
        "internal_ms",
        "kernel_ns_per_elem",
        "kernel_ns_per_unit",
    ]
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def main() -> None:
    ap = argparse.ArgumentParser(description="Import legacy Filip XLSX results and generate comparison plots.")
    ap.add_argument("--xlsx", action="append", required=True, help="Path to legacy XLSX file. Repeat for QSS/SQS/SSQ.")
    ap.add_argument("--optimization-dir", action="append", default=[], help="Optional Filip_original run dir(s) to overlay.")
    ap.add_argument("--legacy-n-qp", type=int, default=0, help="Quadrature points for legacy normalization. 0 = infer.")
    ap.add_argument("--current-operator", default="laplace", help="Operator from current run to compare against legacy reference.")
    ap.add_argument("--out-dir", default="", help="Output directory. Default: analysis/plots/<auto-name>.")
    ap.add_argument("--legacy-label", default="Intel Xe (legacy XLSX)", help="Legend label for legacy series.")
    args = ap.parse_args()

    xlsx_paths = [Path(p).expanduser().resolve() for p in args.xlsx]
    missing = [str(p) for p in xlsx_paths if not p.exists()]
    if missing:
        raise SystemExit(f"Missing XLSX files: {', '.join(missing)}")

    current_dirs = [Path(p).expanduser().resolve() for p in args.optimization_dir]
    for p in current_dirs:
        if not p.exists():
            raise SystemExit(f"Missing optimization dir: {p}")

    n_qp, n_qp_reason = _guess_n_qp(xlsx_paths, int(args.legacy_n_qp))
    case_label = _infer_case_label(xlsx_paths)
    default_out = PLOTS_ROOT / (
        f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}__{_slug(case_label)}"
        + ("__compare" if current_dirs else "__reference")
    )
    out_dir = Path(args.out_dir).expanduser().resolve() if args.out_dir else default_out
    out_dir.mkdir(parents=True, exist_ok=True)

    _setup_plot_style()

    legacy_rows: list[dict[str, Any]] = []
    for path in xlsx_paths:
        legacy_rows.extend(_read_legacy_xlsx(path, n_qp=n_qp, label=args.legacy_label))

    current_runs = [_load_optimization_run(path) for path in current_dirs]
    current_filtered = [_preferred_current_rows(run["rows"], operator=str(args.current_operator).strip().lower()) for run in current_runs]

    legacy_jsonl = out_dir / "legacy_reference_rows.jsonl"
    legacy_csv = out_dir / "legacy_reference_rows.csv"
    _write_rows_jsonl(legacy_rows, legacy_jsonl)
    _write_rows_csv(legacy_rows, legacy_csv)

    # Legacy-only reference: normalized ns/(element*qp)
    fig, axes = plt.subplots(len(VARIANT_ORDER), 1, figsize=(22, 5.0 * len(VARIANT_ORDER) + 1.2), squeeze=False)
    for idx, variant in enumerate(VARIANT_ORDER):
        combos, ys = _best_by_combo(legacy_rows, variant, "kernel_ns_per_unit")
        mapping = {combo: val for combo, val in zip(combos, ys)}
        _plot_series(
            axes[idx][0],
            combos=combos,
            series=[(args.legacy_label, mapping)],
            title=f"{case_label} | {_variant_title(variant)} | kernel ns / (element * qp)",
            ylabel="ns / (element * qp)",
        )
        axes[idx][0].set_xlabel(OPTION_LABEL)
    fig.subplots_adjust(left=0.05, right=0.995, top=0.97, bottom=0.04, hspace=0.56)
    legacy_norm_path = out_dir / "legacy_reference_ns_per_unit.png"
    fig.savefig(legacy_norm_path, dpi=220)
    plt.close(fig)

    # Legacy-only reference: raw kernel ms
    fig, axes = plt.subplots(len(VARIANT_ORDER), 1, figsize=(22, 5.0 * len(VARIANT_ORDER) + 1.2), squeeze=False)
    for idx, variant in enumerate(VARIANT_ORDER):
        combos, ys = _best_by_combo(legacy_rows, variant, "kernel_ms")
        mapping = {combo: val for combo, val in zip(combos, ys)}
        _plot_series(
            axes[idx][0],
            combos=combos,
            series=[(args.legacy_label, mapping)],
            title=f"{case_label} | {_variant_title(variant)} | kernel time",
            ylabel="Kernel time [ms]",
        )
        axes[idx][0].set_xlabel(OPTION_LABEL)
    fig.subplots_adjust(left=0.05, right=0.995, top=0.97, bottom=0.04, hspace=0.56)
    legacy_ms_path = out_dir / "legacy_reference_kernel_ms.png"
    fig.savefig(legacy_ms_path, dpi=220)
    plt.close(fig)

    comparison_plots: list[str] = []
    comparison_notes: list[str] = []
    if current_runs:
        current_labels = [str(run["summary"].get("resolved_backend", run["summary"].get("backend", "run"))) + " | " + str(run["summary"].get("device", "unknown")) for run in current_runs]
        for run in current_runs:
            summary = run["summary"]
            element_types = [str(x).lower() for x in summary.get("element_types", [])]
            operators = [str(x).lower() for x in summary.get("operators", [])]
            if "prism" not in " ".join(element_types):
                comparison_notes.append(
                    f"Current run {summary.get('out_dir', '') or summary.get('backend', '')}: element_types={element_types or ['unknown']} (legacy reference is prism)"
                )
            if str(args.current_operator).strip().lower() not in operators:
                comparison_notes.append(
                    f"Current run {summary.get('out_dir', '') or summary.get('backend', '')}: operator {args.current_operator} not declared in summary operators={operators}"
                )

        fig, axes = plt.subplots(len(VARIANT_ORDER), 1, figsize=(22, 5.2 * len(VARIANT_ORDER) + 1.2), squeeze=False)
        for idx, variant in enumerate(VARIANT_ORDER):
            combos = _union_combos([legacy_rows, *current_filtered], variant)
            legacy_combo, legacy_vals = _best_by_combo(legacy_rows, variant, "kernel_ns_per_unit")
            series = [(args.legacy_label, {c: v for c, v in zip(legacy_combo, legacy_vals)})]
            for label, rows in zip(current_labels, current_filtered):
                cur_combo, cur_vals = _best_by_combo(rows, variant, "kernel_ns_per_unit")
                series.append((label, {c: v for c, v in zip(cur_combo, cur_vals)}))
            _plot_series(
                axes[idx][0],
                combos=combos,
                series=series,
                title=f"Legacy vs current | {_variant_title(variant)} | normalized kernel time",
                ylabel="ns / (element * qp)",
            )
            axes[idx][0].set_xlabel(OPTION_LABEL)
        if comparison_notes:
            fig.text(0.01, 0.995, " | ".join(comparison_notes[:3]), ha="left", va="top", fontsize=8, color="#7c2d12")
        fig.subplots_adjust(left=0.05, right=0.995, top=0.96, bottom=0.04, hspace=0.58)
        cmp_norm_path = out_dir / "legacy_vs_current_ns_per_unit.png"
        fig.savefig(cmp_norm_path, dpi=220)
        plt.close(fig)
        comparison_plots.append(str(cmp_norm_path))

        fig, axes = plt.subplots(len(VARIANT_ORDER), 1, figsize=(22, 5.2 * len(VARIANT_ORDER) + 1.2), squeeze=False)
        for idx, variant in enumerate(VARIANT_ORDER):
            combos = _union_combos([legacy_rows, *current_filtered], variant)
            legacy_combo, legacy_vals = _best_by_combo(legacy_rows, variant, "kernel_ms")
            series = [(args.legacy_label, {c: v for c, v in zip(legacy_combo, legacy_vals)})]
            for label, rows in zip(current_labels, current_filtered):
                cur_combo, cur_vals = _best_by_combo(rows, variant, "kernel_ms")
                series.append((label, {c: v for c, v in zip(cur_combo, cur_vals)}))
            _plot_series(
                axes[idx][0],
                combos=combos,
                series=series,
                title=f"Legacy vs current | {_variant_title(variant)} | raw kernel time",
                ylabel="Kernel time [ms]",
            )
            axes[idx][0].set_xlabel(OPTION_LABEL)
        if comparison_notes:
            fig.text(0.01, 0.995, " | ".join(comparison_notes[:3]), ha="left", va="top", fontsize=8, color="#7c2d12")
        fig.subplots_adjust(left=0.05, right=0.995, top=0.96, bottom=0.04, hspace=0.58)
        cmp_ms_path = out_dir / "legacy_vs_current_kernel_ms.png"
        fig.savefig(cmp_ms_path, dpi=220)
        plt.close(fig)
        comparison_plots.append(str(cmp_ms_path))

    summary = {
        "created_at": datetime.now().astimezone().isoformat(),
        "legacy_files": [str(p) for p in xlsx_paths],
        "legacy_label": args.legacy_label,
        "legacy_case_label": case_label,
        "legacy_n_qp": int(n_qp),
        "legacy_n_qp_reason": n_qp_reason,
        "current_operator": str(args.current_operator).strip().lower(),
        "optimization_dirs": [str(p) for p in current_dirs],
        "legacy_reference_jsonl": str(legacy_jsonl),
        "legacy_reference_csv": str(legacy_csv),
        "generated_plots": [str(legacy_norm_path), str(legacy_ms_path), *comparison_plots],
        "comparison_notes": comparison_notes,
        "out_dir": str(out_dir),
    }
    (out_dir / "summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=True), encoding="utf-8")
    print(json.dumps(summary, indent=2, ensure_ascii=True))


if __name__ == "__main__":
    main()
